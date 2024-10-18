# Import necessary modules
from flask import Flask, render_template, Response
import cv2
import numpy as np
import os
import openpyxl
from datetime import datetime

# Flask app
app = Flask(__name__, static_url_path='/static')

# Video stream URL from phone (you may replace it with your actual stream)
url = 'http://192.0.0.4:8080/video'

# Size reduction factor
size = 2

# Haar cascade file for face detection
classifier = 'haarcascade_frontalface_default.xml'

# Directory where training images are stored
image_dir = 'images'

# Face recognition setup
print("Face Recognition Starting ...")
(images, labels, names, id) = ([], [], {}, 0)

# Load training images
for (subdirs, dirs, files) in os.walk(image_dir):
    for subdir in dirs:
        names[id] = subdir
        subjectpath = os.path.join(image_dir, subdir)
        for filename in os.listdir(subjectpath):
            f_name, f_extension = os.path.splitext(filename)
            if f_extension.lower() not in ['.png', '.jpg', '.jpeg', '.gif', '.pgm']:
                print("Skipping " + filename + ", wrong file type")
                continue
            path = os.path.join(subjectpath, filename)
            label = id
            images.append(cv2.imread(path, 0))
            labels.append(int(label))
        id += 1

# Dimensions for resizing the faces
(im_width, im_height) = (120, 102)

# Convert images and labels to NumPy arrays
(images, labels) = [np.array(lis) for lis in [images, labels]]

# Train the recognizer (using LBPH algorithm)
model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, labels)

# Load the Haar cascade for face detection
haar_cascade = cv2.CascadeClassifier(classifier)

# Initialize Excel workbook and worksheet for logging names and durations
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Face Detection Log"
ws["A1"] = "Name"
ws["B1"] = "Duration (seconds)"

# Dictionary to hold the recognition times
recognition_times = {}
last_recognized_name = None
face_start_time = None

# Function to update the Excel file
def update_excel():
    ws.delete_rows(2, ws.max_row)
    for name, duration in recognition_times.items():
        ws.append([name, duration])
    wb.save("face_detection_log.xlsx")

# Function to generate video frames and perform face recognition
def generate_frames():
    global last_recognized_name, face_start_time
    camera = cv2.VideoCapture(url)

    while True:
        rval, frame = camera.read()
        if not rval:
            print("Failed to open webcam. Trying again...")
            break

        # Flip the frame horizontally
        frame = cv2.flip(frame, 1)

        # Convert the frame to grayscale for face detection
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Resize the frame to speed up detection
        mini = cv2.resize(gray, (int(gray.shape[1] / size), int(gray.shape[0] / size)))

        # Detect faces in the frame
        faces = haar_cascade.detectMultiScale(mini)
        current_time = datetime.now()

        if len(faces) > 0:
            for i in range(len(faces)):
                face_i = faces[i]

                # Get face coordinates and scale them back to the original size
                (x, y, w, h) = [v * size for v in face_i]
                face = gray[y:y + h, x:x + w]
                face_resize = cv2.resize(face, (im_width, im_height))

                # Predict the face using the trained model
                prediction = model.predict(face_resize)
                start = (x, y)
                end = (x + w, y + h)

                # Draw a bounding box around the face
                cv2.rectangle(frame, start, end, (0, 255, 0), 3)
                cv2.rectangle(frame, (start[0], start[1] - 20), (start[0] + 120, start[1]), (0, 255, 255), -3)

                # Add the name or 'Unknown' label
                if prediction[1] < 90:
                    name = names[prediction[0]]
                    cv2.putText(frame, '%s - %.0f' % (name, prediction[1]), (x + 5, y - 5),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)
                    print('%s - %.0f' % (name, prediction[1]))

                    # Handle recognition timing
                    if last_recognized_name == name:
                        # If still the same face, update the duration
                        if face_start_time:
                            duration = (current_time - face_start_time).total_seconds()
                            recognition_times[name] = recognition_times.get(name, 0) + duration
                            face_start_time = current_time  # Reset the start time
                    else:
                        # Log the previous face if it was recognized
                        if last_recognized_name:
                            if last_recognized_name in recognition_times:
                                recognition_times[last_recognized_name] += (current_time - face_start_time).total_seconds()
                        last_recognized_name = name
                        face_start_time = current_time  # Start timing the new face

                        # Ensure the new name is in the recognition_times dictionary
                        if name not in recognition_times:
                            recognition_times[name] = 0  # Initialize if not already present
                            recognition_times[name] += (current_time - face_start_time).total_seconds()

                else:
                    cv2.putText(frame, "Unknown %.0f" % prediction[1], (x + 5, y - 5),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 2)
                    print("Unknown -", prediction[1])

        else:
            # If no faces are detected, log the time of the last recognized face
            if last_recognized_name and face_start_time:
                recognition_times[last_recognized_name] += (current_time - face_start_time).total_seconds()
                last_recognized_name = None
                face_start_time = None

        # Update the Excel file periodically
        update_excel()

        # Encode the frame for web streaming
        ret, buffer = cv2.imencode('.jpg', frame)
        frame = buffer.tobytes()

        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

# Flask routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

# Run the Flask app
if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0')
